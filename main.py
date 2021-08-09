import time

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl


def scrape_comments():
    driver = webdriver.Chrome()
    try:
        try:
            workbook = openpyxl.load_workbook("sample.xlsx")
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            workbook.save("test.xlsx")

        if 'Output' in workbook.sheetnames:
            sheet = workbook["Output"]
        else:
            sheet = workbook.create_sheet("Output", 0)

        sheet.cell(1, 1).value = "#"
        sheet.cell(1, 2).value = "Comment Text"

        i = 1
        for row in sheet.iter_rows():
            if row[1].value is None:
                break
            i += 1

        urls = [
            "https://www.youtube.com/watch?v=N-X8Ztq3jfo",
            "https://www.youtube.com/watch?v=8ZZoxJuxE3A&ab_channel=Ap%C3%A9Amma",
            "https://www.youtube.com/watch?v=Z_BRwG7zNeY&ab_channel=ASHTV",
            "https://www.youtube.com/watch?v=pRzeHSaQJOo&ab_channel=Dr.RohanaWeerasingheOfficial",
            "https://www.youtube.com/watch?v=wpWZqIMY1ek&ab_channel=TechTrack",
            "https://www.youtube.com/watch?v=FidPiSflTKo&ab_channel=SLTechTube"
        ]

        for url in urls:
            print(f"Starting scraping {url}")
            driver.get(url)
            time.sleep(5)

            driver.find_element_by_tag_name('body').send_keys(Keys.PAGE_DOWN)
            last_height = driver.execute_script("return document.documentElement.scrollHeight")

            while True:
                # Wait to load page
                time.sleep(5)
                # Scroll down to bottom
                driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
                new_height = driver.execute_script("return document.documentElement.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height

            ytd_comment_elements = driver.find_elements_by_css_selector("ytd-comment-renderer #content ")

            print("Starting i : ", i)
            for x in ytd_comment_elements:
                t = x.text
                sheet.cell(i, 1).value = i-1
                sheet.cell(i, 2).value = t
                i += 1

            break #todo: remove this line

        workbook.save("sample.xlsx")
        workbook.close()
        driver.close()
        print("Scraping completed successfully")
    except Exception:
        driver.close()


if __name__ == "__main__":
    scrape_comments()
